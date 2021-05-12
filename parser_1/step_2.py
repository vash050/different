import openpyxl as opx

wb = opx.load_workbook('name_bird.xlsx')
sheet = wb.active
for num_row in range(1, sheet.max_row + 1):
    for num_column in range(1, sheet.max_column + 1):
        d = sheet.cell(row=num_row, column=num_column)
        print(d.value)
